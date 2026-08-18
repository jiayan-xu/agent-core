"""Office-basic MCP server for agent-core (A/B dual-track pilot).

Implements the same custom JSON-RPC-over-stdio protocol as the dashboard MCP
server (protocol 2026-07-28, OpenAI-style tools/list). Tools:
  - read_text   read a UTF-8 text file
  - write_text  create/overwrite a UTF-8 text file
  - append_text append a line to a UTF-8 text file
  - excel_group_sum  openpyxl group-by aggregation
"""
import json
import os
import sys
import zipfile
from pathlib import Path

from openpyxl import load_workbook

DENY_COMPONENTS = {".ssh", ".gnupg", ".aws", ".azure", ".config/gcloud"}
DENY_FILENAMES = {
    "id_ed25519", "id_rsa", "id_dsa", "id_ecdsa",
    "id_ed25519.pub", "credentials", "gserviceaccount.json",
}
MAX_XLSX_BYTES = 20 * 1024 * 1024
MAX_XLSX_UNCOMPRESSED = 100 * 1024 * 1024
MAX_XLSX_FILES = 1000
MAX_XLSX_ROWS = 100_000
MAX_GROUPS = 10_000


def check_xlsx_zip(source):
    """Reject zip bombs and oversized workbooks before openpyxl decompresses them."""
    try:
        with zipfile.ZipFile(source) as zf:
            infos = zf.infolist()
            if len(infos) > MAX_XLSX_FILES:
                raise ValueError(f"xlsx has too many entries ({len(infos)} > {MAX_XLSX_FILES})")
            total = sum(info.file_size for info in infos)
            if total > MAX_XLSX_UNCOMPRESSED:
                raise ValueError(f"xlsx uncompressed size too large ({total} > {MAX_XLSX_UNCOMPRESSED})")
    except zipfile.BadZipFile as e:
        raise ValueError(f"not a valid xlsx: {e}")


def safe_path(raw: str, require_root: bool = False) -> Path:
    """Reject sensitive paths and (when required) paths outside the sandbox root."""
    p = Path(raw).expanduser()
    if not p.is_absolute():
        raise ValueError(f"path must be absolute: {raw}")
    root = os.environ.get("AGENT_SANDBOX_ROOT")
    # 写类工具必须 fail-closed：没有沙箱根就拒绝启动，不能退化成任意路径写。
    if require_root and not root:
        raise RuntimeError("AGENT_SANDBOX_ROOT is required for file-writing tools")
    # 先做符号链接解析（与 Rust 侧 normalize 一致），再做敏感段匹配，
    # 否则 /tmp/link -> ~/.ssh 这类链接路径会绕过 deny 列表。
    p = p.resolve(strict=False)
    parts = [part.lower() for part in p.parts if part]
    for comp in DENY_COMPONENTS:
        segs = comp.split("/")
        if any(parts[i:i + len(segs)] == segs for i in range(len(parts) - len(segs) + 1)):
            raise ValueError(f"path hits a sensitive directory: {raw}")
    if p.name.lower() in DENY_FILENAMES or p.name.lower().endswith((".pem", ".key")):
        raise ValueError(f"path hits a sensitive file: {raw}")
    if root:
        try:
            root_resolved = Path(root).resolve(strict=False)
            if not p.is_relative_to(root_resolved):
                raise ValueError(f"path is outside sandbox root {root}: {raw}")
        except OSError:
            raise
    return p


def _flags(*names):
    value = 0
    for name in names:
        value |= getattr(os, name, 0)
    return value


def open_text(path: Path, mode: str):
    """Open a checked path with O_NOFOLLOW and re-verify the opened fd.

    Fails closed on platforms without O_NOFOLLOW (no final-component symlink
    protection). On Linux the opened fd is re-resolved and re-checked so an
    intermediate directory swapped after safe_path() cannot redirect I/O.
    """
    if not hasattr(os, "O_NOFOLLOW"):
        raise RuntimeError("platform lacks O_NOFOLLOW; refusing to open files")
    if mode == "r":
        flags = _flags("O_RDONLY", "O_NOFOLLOW", "O_CLOEXEC")
    elif mode == "w":
        flags = _flags("O_WRONLY", "O_CREAT", "O_TRUNC", "O_NOFOLLOW", "O_CLOEXEC")
    elif mode == "a":
        flags = _flags("O_WRONLY", "O_CREAT", "O_APPEND", "O_NOFOLLOW", "O_CLOEXEC")
    else:
        raise ValueError(f"unsupported mode {mode}")
    fd = os.open(str(path), flags, 0o666)
    try:
        if sys.platform.startswith("linux"):
            fd_path = Path(f"/proc/self/fd/{fd}").resolve(strict=False)
            fd_parts = [part.lower() for part in fd_path.parts if part]
            for comp in DENY_COMPONENTS:
                segs = comp.split("/")
                if any(fd_parts[i:i + len(segs)] == segs for i in range(len(fd_parts) - len(segs) + 1)):
                    raise ValueError(f"opened path resolves into a sensitive directory: {fd_path}")
            root = os.environ.get("AGENT_SANDBOX_ROOT")
            if root and not fd_path.is_relative_to(Path(root).resolve(strict=False)):
                raise ValueError(f"opened path escapes sandbox root {root}: {fd_path}")
        return open(fd, mode, encoding="utf-8", errors="replace", closefd=True)
    except Exception:
        os.close(fd)
        raise

def open_binary_nofollow(path: Path):
    """Open a checked path in binary mode with O_NOFOLLOW for openpyxl."""
    if not hasattr(os, "O_NOFOLLOW"):
        raise RuntimeError("platform lacks O_NOFOLLOW; refusing to open xlsx files")
    flags = _flags("O_RDONLY", "O_NOFOLLOW", "O_CLOEXEC")
    fd = os.open(str(path), flags)
    try:
        if sys.platform.startswith("linux"):
            fd_path = Path(f"/proc/self/fd/{fd}").resolve(strict=False)
            fd_parts = [part.lower() for part in fd_path.parts if part]
            for comp in DENY_COMPONENTS:
                segs = comp.split("/")
                if any(fd_parts[i:i + len(segs)] == segs for i in range(len(fd_parts) - len(segs) + 1)):
                    raise ValueError(f"opened path resolves into a sensitive directory: {fd_path}")
            root = os.environ.get("AGENT_SANDBOX_ROOT")
            if root and not fd_path.is_relative_to(Path(root).resolve(strict=False)):
                raise ValueError(f"opened path escapes sandbox root {root}: {fd_path}")
        return os.fdopen(fd, "rb")
    except Exception:
        os.close(fd)
        raise


TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "read_text",
            "description": "读取一个 UTF-8 文本文件并返回内容。",
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {"type": "string", "description": "文件绝对路径"},
                    "max_chars": {"type": "number", "description": "最大返回字符数，默认 8000"},
                },
                "required": ["path"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "write_text",
            "description": "创建或覆盖一个 UTF-8 文本文件。",
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {"type": "string", "description": "文件绝对路径"},
                    "content": {"type": "string", "description": "完整写入内容"},
                },
                "required": ["path", "content"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "append_text",
            "description": "向 UTF-8 文本文件追加一行。",
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {"type": "string", "description": "文件绝对路径"},
                    "content": {"type": "string", "description": "追加内容"},
                },
                "required": ["path", "content"],
            },
        },
    },
    {
        "type": "function",
        "function": {
            "name": "excel_group_sum",
            "description": "读取 xlsx 并按指定列分组求和，返回从高到低的排名。",
            "parameters": {
                "type": "object",
                "properties": {
                    "path": {"type": "string", "description": "xlsx 文件绝对路径"},
                    "group_col": {"type": "string", "description": "分组列名，默认 product"},
                    "value_col": {"type": "string", "description": "求和列名，默认 qty"},
                },
                "required": ["path"],
            },
        },
    },
]


def handle(req):
    if not isinstance(req, dict):
        return {"jsonrpc": "2.0", "id": None,
                "error": {"code": -32600, "message": "invalid request: expected a JSON object"}}
    rid = req.get("id", 1)
    method = req.get("method", "")
    raw_params = req.get("params")
    params = raw_params if isinstance(raw_params, dict) else {}
    if method == "initialize":
        return {"jsonrpc": "2.0", "id": rid, "result": {
            "protocolVersion": "2026-07-28", "capabilities": {"tools": {}},
            "serverInfo": {"name": "office-basic", "version": "0.1"}}}
    if method == "tools/list":
        return {"jsonrpc": "2.0", "id": rid, "result": {"tools": TOOLS}}
    if method == "tools/call":
        name = params.get("name")
        args = params.get("arguments", {}) or {}
        try:
            if name == "read_text":
                path = safe_path(args["path"])
                cap = min(1_000_000, max(0, int(args.get("max_chars", 8000))))
                # 只读取 cap+64 字符，cap 同时约束 I/O 与返回内容，避免大文件整读进内存。
                with open_text(path, "r") as f:
                    text = f.read(cap + 64)
                if len(text) > cap:
                    text = text[:cap] + "\n...[truncated]"
                return {"jsonrpc": "2.0", "id": rid, "result": {
                    "content": [{"type": "text", "text": text}]}}
            if name == "write_text":
                path = safe_path(args["path"], require_root=True)
                path.parent.mkdir(parents=True, exist_ok=True)
                with open_text(path, "w") as f:
                    f.write(args["content"])
                return {"jsonrpc": "2.0", "id": rid, "result": {
                    "content": [{"type": "text", "text": json.dumps({"ok": True, "path": str(path)}, ensure_ascii=False)}]}}
            if name == "append_text":
                path = safe_path(args["path"], require_root=True)
                with open_text(path, "a") as f:
                    f.write(args["content"] + "\n")
                return {"jsonrpc": "2.0", "id": rid, "result": {
                    "content": [{"type": "text", "text": json.dumps({"ok": True, "path": str(path)}, ensure_ascii=False)}]}}
            if name == "excel_group_sum":
                if not hasattr(os, "O_NOFOLLOW"):
                    raise RuntimeError("platform lacks O_NOFOLLOW; refusing to open xlsx files")
                path = safe_path(args["path"])
                fh = open_binary_nofollow(path)
                try:
                    if os.fstat(fh.fileno()).st_size > MAX_XLSX_BYTES:
                        return {"jsonrpc": "2.0", "id": rid, "error": {
                            "code": -32602, "message": f"xlsx too large (max {MAX_XLSX_BYTES} bytes)"}}
                    check_xlsx_zip(fh)
                    fh.seek(0)
                    wb = load_workbook(fh, read_only=True, data_only=True)
                    try:
                        ws = wb.active
                        rows = ws.iter_rows(values_only=True)
                        header_row = next(rows, None)
                        if header_row is None:
                            return {"jsonrpc": "2.0", "id": rid, "result": {
                                "content": [{"type": "text", "text": json.dumps({"ranking": []}, ensure_ascii=False)}]}}
                        header = [str(c) for c in header_row]
                        gc = args.get("group_col", "product")
                        vc = args.get("value_col", "qty")
                        try:
                            gi = header.index(gc)
                            vi = header.index(vc)
                        except ValueError as e:
                            return {"jsonrpc": "2.0", "id": rid,
                                    "error": {"code": -32602, "message": f"invalid params: {e}"}}
                        agg = {}
                        row_count = 0
                        for row in rows:
                            row_count += 1
                            if row_count > MAX_XLSX_ROWS:
                                return {"jsonrpc": "2.0", "id": rid, "error": {
                                    "code": -32602, "message": f"xlsx has too many rows (> {MAX_XLSX_ROWS})"}}
                            key = str(row[gi])
                            raw_val = row[vi]
                            if raw_val is None or isinstance(raw_val, bool):
                                continue
                            try:
                                val = float(raw_val)
                            except (TypeError, ValueError):
                                # 非数字单元格不静默当 0，跳过并继续聚合可解释的数据
                                continue
                            if key not in agg and len(agg) >= MAX_GROUPS:
                                return {"jsonrpc": "2.0", "id": rid, "error": {
                                    "code": -32602, "message": f"too many distinct groups (> {MAX_GROUPS})"}}
                            agg[key] = agg.get(key, 0) + val
                        ranking = [{"group": k, "total": v} for k, v in
                                   sorted(agg.items(), key=lambda kv: kv[1], reverse=True)]
                        return {"jsonrpc": "2.0", "id": rid, "result": {
                            "content": [{"type": "text", "text": json.dumps({"ranking": ranking}, ensure_ascii=False)}]}}
                    finally:
                        wb.close()
                finally:
                    fh.close()
            return {"jsonrpc": "2.0", "id": rid, "error": {"code": -32602, "message": f"unknown tool {name}"}}
        except (KeyError, IndexError, ValueError, TypeError) as e:
            # 客户端参数错误（缺参数 / 非整数 / 非法路径 / 列不存在 / 类型错误）→ -32602
            return {"jsonrpc": "2.0", "id": rid,
                    "error": {"code": -32602, "message": f"invalid params: {e}"}}
        except Exception as e:
            # 只记日志，不回显内部细节；给调用方一个通用 -32603。
            print(f"[office-basic] internal error: {e}", file=sys.stderr, flush=True)
            return {"jsonrpc": "2.0", "id": rid,
                    "error": {"code": -32603, "message": "internal error"}}
    return {"jsonrpc": "2.0", "id": rid, "error": {"code": -32601, "message": f"unknown method {method}"}}


def main():
    if hasattr(sys.stdin, "reconfigure"):
        sys.stdin.reconfigure(encoding="utf-8", errors="replace")
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")
    print(json.dumps({"jsonrpc": "2.0", "type": "ready", "tools": len(TOOLS)}), flush=True)
    for line in sys.stdin:
        line = line.strip()
        if not line:
            continue
        try:
            req = json.loads(line)
        except json.JSONDecodeError as e:
            print(json.dumps({"jsonrpc": "2.0", "id": None,
                              "error": {"code": -32700, "message": f"parse error: {e}"}},
                             ensure_ascii=False), flush=True)
            continue
        resp = handle(req)
        print(json.dumps(resp, ensure_ascii=False), flush=True)


if __name__ == "__main__":
    main()
